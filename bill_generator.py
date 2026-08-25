import os
import shutil
from datetime import datetime

from openpyxl import load_workbook
from dotenv import load_dotenv


load_dotenv()


BILL_TEMPLATE_PATH = os.getenv(
    "BILL_TEMPLATE_PATH",
    "templates/FORM_BILL.xlsx"
)

OUTPUT_PATH = os.getenv(
    "OUTPUT_PATH",
    "output"
)


def _write_multiline(sheet, start_row, column, text, max_lines):
    """
    Ghi text nhiều dòng xuống nhiều ô liên tiếp.

    Ví dụ:
    B6
    B7
    B8...
    """

    if not text:
        text = ""

    lines = [
        line.strip()
        for line in str(text).splitlines()
        if line.strip()
    ]

    # Nếu user nhập một dòng thì vẫn ghi được.
    if not lines:
        lines = [""]

    for index in range(max_lines):

        value = ""

        if index < len(lines):
            value = lines[index]

        sheet.cell(
            row=start_row + index,
            column=column,
            value=value
        )


def generate_bill(
    booking,
    output_file=None
):
    """
    Tạo Bill mới từ FORM_BILL.xlsx
    dựa trên một row booking trong Excel database.
    """

    if not booking:
        raise ValueError(
            "Booking data is empty."
        )

    if not os.path.exists(
        BILL_TEMPLATE_PATH
    ):
        raise FileNotFoundError(
            f"Không tìm thấy template: {BILL_TEMPLATE_PATH}"
        )

    os.makedirs(
        OUTPUT_PATH,
        exist_ok=True
    )

    booking_id = (
        booking.get("id")
        or "BOOKING"
    )

    if output_file is None:

        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        output_file = os.path.join(
            OUTPUT_PATH,
            f"BILL_{booking_id}_{timestamp}.xlsx"
        )

    # Copy template để không sửa file gốc
    shutil.copy2(
        BILL_TEMPLATE_PATH,
        output_file
    )

    workbook = load_workbook(
        output_file
    )

    sheet_name = None

    for name in workbook.sheetnames:
        if name.strip() == "BL FORM":
            sheet_name = name
            break

    if sheet_name is None:

        workbook.close()

        raise ValueError(
            f"Template không có sheet BL FORM. "
            f"Các sheet hiện có: {workbook.sheetnames}"
        )

    sheet = workbook[
        sheet_name
    ]


    # =====================================================
    # SHIPPER
    # BL FORM: B6:B10
    # =====================================================

    _write_multiline(
        sheet,
        start_row=6,
        column=2,
        text=booking.get(
            "shipper",
            ""
        ),
        max_lines=5
    )


    # =====================================================
    # CONSIGNEE
    # BL FORM: B13:B18
    # =====================================================

    _write_multiline(
        sheet,
        start_row=13,
        column=2,
        text=booking.get(
            "consignee",
            ""
        ),
        max_lines=6
    )


    # =====================================================
    # NOTIFY PARTY
    # =====================================================

    notify_party = booking.get(
        "notify_party",
        ""
    )

    if notify_party:

        sheet["A20"] = notify_party


    # =====================================================
    # BOOKING / B/L NUMBER
    #
    # Template hiện đặt B/L No. tại M8.
    # Tạm dùng booking_no làm reference.
    # Sau này nếu database có BL No riêng thì đổi mapping.
    # =====================================================

    sheet["M8"] = (
        booking.get(
            "booking_no",
            ""
        )
    )


    # =====================================================
    # ROUTING
    # =====================================================

    sheet["E27"] = (
        booking.get(
            "port_of_receipt",
            ""
        )
    )

    sheet["A29"] = (
        booking.get(
            "vessel_voyage",
            ""
        )
    )

    sheet["E29"] = (
        booking.get(
            "port_of_loading",
            ""
        )
    )

    sheet["K29"] = (
        booking.get(
            "port_of_discharge",
            ""
        )
    )

    # Port of delivery
    # hiện tạm sử dụng POD nếu chưa có field riêng.
    sheet["O29"] = (
        booking.get(
            "port_of_discharge",
            ""
        )
    )


    # =====================================================
    # PACKAGE QUANTITY
    # =====================================================

    package_quantity = booking.get(
        "package_quantity",
        ""
    )

    package_type = booking.get(
        "package_type",
        ""
    )

    package_text = ""

    if package_quantity:
        package_text += str(
            package_quantity
        )

    if package_type:

        if package_text:
            package_text += " "

        package_text += str(
            package_type
        )

    if package_text:
        sheet["F33"] = package_text


    # =====================================================
    # PRODUCT DESCRIPTION
    # =====================================================

    product_name = (
        booking.get(
            "product_name",
            ""
        )
        or ""
    )

    product_description = (
        booking.get(
            "product_description",
            ""
        )
        or ""
    )

    proper_shipping_name = (
        booking.get(
            "proper_shipping_name",
            ""
        )
        or ""
    )

    description_lines = []

    if product_name:
        description_lines.append(
            product_name
        )

    if proper_shipping_name:
        description_lines.append(
            proper_shipping_name
        )

    if product_description:
        description_lines.append(
            product_description
        )

    description = "\n".join(
        description_lines
    )

    # BL FORM description area starts around I36
    _write_multiline(
        sheet,
        start_row=36,
        column=9,
        text=description,
        max_lines=8
    )


    # =====================================================
    # HS CODE
    # =====================================================

    hs_code = booking.get(
        "hs_code",
        ""
    )

    if hs_code:

        sheet["I47"] = (
            f"HS CODE: {hs_code}"
        )


    # =====================================================
    # GROSS WEIGHT
    # =====================================================

    cargo_weight = booking.get(
        "cargo_weight"
    )

    if cargo_weight not in [
        None,
        ""
    ]:

        try:

            weight = float(
                cargo_weight
            )

            sheet["P33"] = (
                f"{weight:,.2f} KGS"
            )

        except (
            ValueError,
            TypeError
        ):

            sheet["P33"] = str(
                cargo_weight
            )


    # =====================================================
    # CBM / MEASUREMENT
    # =====================================================

    cargo_volume = booking.get(
        "cargo_volume"
    )

    if cargo_volume not in [
        None,
        ""
    ]:

        try:

            volume = float(
                cargo_volume
            )

            sheet["Q33"] = (
                f"{volume:g} CBM"
            )

        except (
            ValueError,
            TypeError
        ):

            sheet["Q33"] = str(
                cargo_volume
            )


    # =====================================================
    # CONTAINER SUMMARY
    # =====================================================

    container_quantity = (
        booking.get(
            "container_quantity",
            ""
        )
        or ""
    )

    if container_quantity:

        sheet["A38"] = (
            f"TTL: {container_quantity}"
        )


    # =====================================================
    # MARKS & NUMBERS
    # =====================================================

    marks_numbers = (
        booking.get(
            "marks_numbers",
            ""
        )
        or ""
    )

    if marks_numbers:

        _write_multiline(
            sheet,
            start_row=41,
            column=1,
            text=marks_numbers,
            max_lines=8
        )


    # =====================================================
    # IMO
    # =====================================================

    imo_details = (
        booking.get(
            "imo_details",
            ""
        )
        or ""
    )

    if imo_details:

        sheet["I48"] = (
            imo_details
        )


    # =====================================================
    # SAVE
    # =====================================================

    workbook.save(
        output_file
    )

    workbook.close()

    return output_file