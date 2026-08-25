import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv


load_dotenv()


EXCEL_DB_PATH = os.getenv(
    "BOOKING_DATABASE_PATH",
    "data/BOOKING_DATABASE.xlsx"
)

SHEET_NAME = "BOOKINGS"


HEADERS = [
    "ID",
    "Created At",
    "Updated At",
    "Status",

    "Customer Name",
    "Customer Email",

    "Carrier",
    "Booking Account",

    "Product Name",
    "Product Description",
    "HS Code",

    "Cargo Weight KG",
    "Cargo Volume CBM",

    "Container Quantity",
    "Package Quantity",
    "Package Type",

    "Port of Receipt",
    "Port of Loading",
    "Port of Discharge",

    "Booking No",
    "Vessel Voyage",

    "CY Cut Off",
    "VGM Cut Off",
    "SI Cut Off",

    "POL ETA",
    "ETD",
    "POD ETA",

    "GWT + Tare",
    "Limit GWT Each",

    "IMO Class",
    "UN No",
    "Proper Shipping Name",

    "Empty Pick Up",
    "Inland Return",

    "Shipper",
    "Consignee",
    "Notify Party",
    "Marks Numbers",

    "Source Email",
    "Confirmation PDF",
    "Bill File"
]


FIELD_TO_COLUMN = {
    "id": "ID",
    "created_at": "Created At",
    "updated_at": "Updated At",
    "status": "Status",

    "customer_name": "Customer Name",
    "customer_email": "Customer Email",

    "carrier_name": "Carrier",
    "booking_account": "Booking Account",

    "product_name": "Product Name",
    "product_description": "Product Description",
    "hs_code": "HS Code",

    "cargo_weight": "Cargo Weight KG",
    "cargo_volume": "Cargo Volume CBM",

    "container_quantity": "Container Quantity",
    "package_quantity": "Package Quantity",
    "package_type": "Package Type",

    "port_of_receipt": "Port of Receipt",
    "port_of_loading": "Port of Loading",
    "port_of_discharge": "Port of Discharge",

    "booking_no": "Booking No",
    "vessel_voyage": "Vessel Voyage",

    "cy_cut_off": "CY Cut Off",
    "vgm_cut_off": "VGM Cut Off",
    "si_cut_off": "SI Cut Off",

    "pol_eta_date": "POL ETA",
    "etd_date": "ETD",
    "pod_eta_date": "POD ETA",
    "eta_date": "POD ETA",

    "gwt_plus_tare": "GWT + Tare",
    "limit_gwt_each": "Limit GWT Each",

    "imo_class": "IMO Class",
    "un_no": "UN No",
    "proper_shipping_name": "Proper Shipping Name",

    "empty_pickup": "Empty Pick Up",
    "inland_return": "Inland Return",

    "shipper": "Shipper",
    "consignee": "Consignee",
    "notify_party": "Notify Party",
    "marks_numbers": "Marks Numbers",

    "source_email": "Source Email",
    "source_pdf": "Confirmation PDF",
    "bill_file": "Bill File"
}


def init_db():
    """
    Tạo BOOKING_DATABASE.xlsx nếu chưa tồn tại.
    """

    os.makedirs(
        os.path.dirname(EXCEL_DB_PATH),
        exist_ok=True
    )

    if os.path.exists(EXCEL_DB_PATH):
        return

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = SHEET_NAME

    for col_index, header in enumerate(
        HEADERS,
        start=1
    ):
        cell = sheet.cell(
            row=1,
            column=col_index,
            value=header
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    sheet.freeze_panes = "A2"

    widths = {
        "A": 14,
        "B": 20,
        "C": 20,
        "D": 14,
        "E": 28,
        "F": 30,
        "G": 20,
        "H": 18,
        "I": 32,
        "J": 40,
        "K": 16,
        "L": 18,
        "M": 18,
        "N": 22,
        "Q": 24,
        "R": 24,
        "S": 26,
        "T": 20,
        "U": 28,
        "V": 22,
        "W": 22,
        "X": 22,
        "Y": 18,
        "Z": 18,
        "AA": 18,
        "AB": 22,
        "AC": 20,
        "AD": 14,
        "AE": 14,
        "AF": 40,
        "AG": 35,
        "AH": 40,
        "AI": 35,
        "AJ": 35,
        "AK": 35,
        "AL": 30,
        "AM": 40,
        "AN": 30,
        "AO": 30
    }

    for column, width in widths.items():
        sheet.column_dimensions[
            column
        ].width = width

    workbook.save(
        EXCEL_DB_PATH
    )


def _load():
    init_db()

    workbook = load_workbook(
        EXCEL_DB_PATH
    )

    sheet = workbook[
        SHEET_NAME
    ]

    return workbook, sheet


def _header_map(sheet):
    """
    {'ID': 1, 'Created At': 2, ...}
    """

    return {
        cell.value: cell.column
        for cell in sheet[1]
        if cell.value
    }


def _generate_booking_id(sheet):
    """
    BK000001, BK000002, ...
    """

    max_number = 0

    for row in range(
        2,
        sheet.max_row + 1
    ):

        value = sheet.cell(
            row=row,
            column=1
        ).value

        if not value:
            continue

        text = str(value)

        if text.startswith("BK"):

            try:
                number = int(
                    text[2:]
                )

                max_number = max(
                    max_number,
                    number
                )

            except ValueError:
                pass

    return (
        f"BK{max_number + 1:06d}"
    )


def insert_booking(data):
    """
    Thêm booking mới vào Excel.
    """

    workbook, sheet = _load()

    header_map = _header_map(
        sheet
    )

    booking_id = (
        _generate_booking_id(
            sheet
        )
    )

    now = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    row_number = (
        sheet.max_row + 1
    )

    base_data = {
        "id": booking_id,
        "created_at": now,
        "updated_at": now,
        "status": data.get(
            "status",
            "NEW"
        )
    }

    merged = {
        **data,
        **base_data
    }

    for field, value in merged.items():

        column_name = (
            FIELD_TO_COLUMN.get(
                field
            )
        )

        if not column_name:
            continue

        column_index = (
            header_map.get(
                column_name
            )
        )

        if not column_index:
            continue

        sheet.cell(
            row=row_number,
            column=column_index,
            value=value
        )

    workbook.save(
        EXCEL_DB_PATH
    )

    return booking_id


def get_all_bookings():
    """
    Đọc toàn bộ booking từ Excel.
    """

    workbook, sheet = _load()

    header_map = _header_map(
        sheet
    )

    reverse_map = {
        excel_header: field
        for field, excel_header
        in FIELD_TO_COLUMN.items()
    }

    result = []

    for row in range(
        2,
        sheet.max_row + 1
    ):

        if not sheet.cell(
            row=row,
            column=1
        ).value:
            continue

        booking = {}

        for header, column in (
            header_map.items()
        ):

            field = reverse_map.get(
                header
            )

            if not field:
                continue

            booking[field] = (
                sheet.cell(
                    row=row,
                    column=column
                ).value
            )

        result.append(
            booking
        )

    workbook.close()

    return list(
        reversed(result)
    )


def get_booking(booking_id):
    """
    Tìm 1 booking theo BKxxxxxx.
    """

    bookings = (
        get_all_bookings()
    )

    for booking in bookings:

        if str(
            booking.get("id")
        ) == str(booking_id):

            return booking

    return None


def update_booking(
    booking_id,
    data
):
    """
    Update đúng row theo Booking ID.
    """

    workbook, sheet = _load()

    header_map = _header_map(
        sheet
    )

    id_column = header_map[
        "ID"
    ]

    target_row = None

    for row in range(
        2,
        sheet.max_row + 1
    ):

        value = sheet.cell(
            row=row,
            column=id_column
        ).value

        if str(value) == str(
            booking_id
        ):

            target_row = row
            break

    if target_row is None:

        workbook.close()
        return False

    for field, value in data.items():

        column_name = (
            FIELD_TO_COLUMN.get(
                field
            )
        )

        if not column_name:
            continue

        column_index = (
            header_map.get(
                column_name
            )
        )

        if not column_index:
            continue

        sheet.cell(
            row=target_row,
            column=column_index,
            value=value
        )

    sheet.cell(
        row=target_row,
        column=header_map[
            "Updated At"
        ],
        value=datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    )

    workbook.save(
        EXCEL_DB_PATH
    )

    return True


def delete_booking(
    booking_id
):
    """
    Xóa row booking khỏi Excel.
    """

    workbook, sheet = _load()

    header_map = _header_map(
        sheet
    )

    id_column = header_map[
        "ID"
    ]

    target_row = None

    for row in range(
        2,
        sheet.max_row + 1
    ):

        if str(
            sheet.cell(
                row=row,
                column=id_column
            ).value
        ) == str(booking_id):

            target_row = row
            break

    if target_row is None:

        workbook.close()
        return False

    sheet.delete_rows(
        target_row,
        1
    )

    workbook.save(
        EXCEL_DB_PATH
    )

    return True


if __name__ == "__main__":

    init_db()

    print(
        f"Excel database ready: {EXCEL_DB_PATH}"
    )